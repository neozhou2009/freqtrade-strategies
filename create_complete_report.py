#!/usr/bin/env python3
"""Create comprehensive Excel report from Freqtrade backtest results"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Styles
title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=9)
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Load data
backtest_dir = '/home/neozh/freqtrade-strategies/user_data/backtest_results/backtest_export'
json_file = f'{backtest_dir}/backtest-result-2026-03-09_17-36-29.json'

with open(json_file, 'r') as f:
    data = json.load(f)

# Parse strategy data
strategy_data = data.get('strategy', {})
nostalgia_data = strategy_data.get('Nostalgia', {})
trades = nostalgia_data.get('trades', [])

print(f"Loaded {len(trades)} trades")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Summary"

# ============= Sheet 1: Summary =============
ws['A1'] = 'Freqtrade Backtest Report - Nostalgia Strategy'
ws['A1'].font = Font(name='Arial', size=16, bold=True)
ws.merge_cells('A1:F1')

# Key metrics from JSON
metrics = nostalgia_data.get('results_per_pair', [])
total_metrics = [m for m in metrics if m.get('key') == 'TOTAL'][0] if metrics else {}

ws['A3'] = 'Backtest Period:'
ws['B3'] = f"{nostalgia_data.get('backtest_start', '')} to {nostalgia_data.get('backtest_end', '')}"
ws['A4'] = 'Strategy:'
ws['B4'] = 'Nostalgia'
ws['A5'] = 'Timeframe:'
ws['B5'] = '5m'
ws['A6'] = 'Exchange:'
ws['B6'] = 'Binance (Futures)'
ws['A7'] = 'Initial Balance:'
ws['B7'] = f"{nostalgia_data.get('starting_balance', 1000)} USDT"
ws['A8'] = 'Final Balance:'
ws['B8'] = f"{nostalgia_data.get('final_balance', 0):.2f} USDT"

# Key metrics
key_metrics = [
    ('Total Trades', total_metrics.get('trades', 0)),
    ('Wins', total_metrics.get('wins', 0)),
    ('Losses', total_metrics.get('losses', 0)),
    ('Draws', total_metrics.get('draws', 0)),
    ('Win Rate', f"{total_metrics.get('winrate', 0)*100:.1f}%"),
    ('Total Profit', f"{total_metrics.get('profit_total_abs', 0):.2f} USDT"),
    ('Profit Factor', f"{total_metrics.get('profit_factor', 0):.2f}"),
    ('CAGR', f"{total_metrics.get('cagr', 0)*100:.2f}%"),
    ('Sharpe', f"{total_metrics.get('sharpe', 0):.2f}"),
    ('Sortino', f"{total_metrics.get('sortino', 0):.2f}"),
    ('Calmar', f"{total_metrics.get('calmar', 0):.2f}"),
    ('SQN', f"{total_metrics.get('sqn', 0):.2f}"),
    ('Max Drawdown', f"{total_metrics.get('max_drawdown_abs', 0):.2f} USDT"),
    ('Market Change', f"{nostalgia_data.get('market_change', 0)*100:.2f}%"),
]

ws['A10'] = 'Key Metrics'
ws['A10'].font = header_font
ws['A10'].fill = header_fill
ws.merge_cells('A10:B10')

for i, (label, value) in enumerate(key_metrics, 11):
    ws[f'A{i}'] = label
    ws[f'B{i}'] = str(value)

# ============= Sheet 2: All Trades =============
ws2 = wb.create_sheet(title="All Trades")

# Headers
trade_headers = [
    'Trade ID', 'Pair', 'Open Date', 'Close Date', 'Direction', 
    'Entry Price', 'Exit Price', 'Amount', 'Profit (USDT)', 'Profit %', 
    'Hold Duration', 'Exit Reason', 'Fee'
]
for col, header in enumerate(trade_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Trade data
for row, trade in enumerate(trades, 2):
    try:
        ws2.cell(row=row, column=1, value=trade.get('trade_id', row-2)).border = thin_border
        ws2.cell(row=row, column=2, value=trade.get('pair', '')).border = thin_border
        
        # Dates
        open_date = trade.get('open_date', '')
        close_date = trade.get('close_date', '')
        ws2.cell(row=row, column=3, value=open_date[:19] if open_date else '').border = thin_border
        ws2.cell(row=row, column=4, value=close_date[:19] if close_date else '').border = thin_border
        
        # Direction
        ws2.cell(row=row, column=5, value='long' if not trade.get('is_short') else 'short').border = thin_border
        
        # Prices
        ws2.cell(row=row, column=6, value=round(trade.get('open_rate', 0), 4)).border = thin_border
        ws2.cell(row=row, column=7, value=round(trade.get('close_rate', 0), 4)).border = thin_border
        ws2.cell(row=row, column=8, value=round(trade.get('amount', 0), 4)).border = thin_border
        
        # Profit
        profit_abs = trade.get('profit_abs', 0)
        profit_ratio = trade.get('profit_ratio', 0)
        profit_cell = ws2.cell(row=row, column=9, value=round(profit_abs, 2))
        profit_cell.border = thin_border
        if profit_abs > 0:
            profit_cell.fill = green_fill
        elif profit_abs < 0:
            profit_cell.fill = red_fill
        
        profit_pct_cell = ws2.cell(row=row, column=10, value=f"{profit_ratio*100:.2f}%")
        profit_pct_cell.border = thin_border
        if profit_ratio > 0:
            profit_pct_cell.fill = green_fill
        elif profit_ratio < 0:
            profit_pct_cell.fill = red_fill
        
        # Duration
        duration_mins = trade.get('trade_duration', 0)
        hours = duration_mins // 60
        mins = duration_mins % 60
        ws2.cell(row=row, column=11, value=f"{hours}h {mins}m").border = thin_border
        
        # Exit Reason
        ws2.cell(row=row, column=12, value=trade.get('exit_reason', '')).border = thin_border
        
        # Fee
        fee = trade.get('fee_open', 0) + trade.get('fee_close', 0)
        ws2.cell(row=row, column=13, value=round(fee, 4)).border = thin_border
        
    except Exception as e:
        print(f"Error processing trade {row}: {e}")

# Adjust column widths
for column in ws2.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws2.column_dimensions[column_letter].width = min(max_length + 2, 25)

# ============= Sheet 3: Pair Analysis =============
ws3 = wb.create_sheet(title="Pair Analysis")

ws3['A1'] = 'Pair Performance Analysis'
ws3['A1'].font = title_font
ws3.merge_cells('A1:J1')

pair_results = [m for m in metrics if m.get('key') != 'TOTAL'] if metrics else []

headers = ['Pair', 'Trades', 'Wins', 'Losses', 'Win Rate', 'Total Profit', 'Avg Profit %', 'Profit Factor', 'Sharpe', 'Calmar']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 4
for pair in pair_results:
    ws3.cell(row=row, column=1, value=pair.get('key', '')).border = thin_border
    ws3.cell(row=row, column=2, value=pair.get('trades', 0)).border = thin_border
    ws3.cell(row=row, column=3, value=pair.get('wins', 0)).border = thin_border
    ws3.cell(row=row, column=4, value=pair.get('losses', 0)).border = thin_border
    ws3.cell(row=row, column=5, value=f"{pair.get('winrate', 0)*100:.1f}%").border = thin_border
    
    profit = pair.get('profit_total_abs', 0)
    profit_cell = ws3.cell(row=row, column=6, value=round(profit, 2))
    profit_cell.border = thin_border
    if profit > 0:
        profit_cell.fill = green_fill
    elif profit < 0:
        profit_cell.fill = red_fill
    
    ws3.cell(row=row, column=7, value=f"{pair.get('profit_mean_pct', 0):.2f}%").border = thin_border
    ws3.cell(row=row, column=8, value=f"{pair.get('profit_factor', 0):.2f}").border = thin_border
    ws3.cell(row=row, column=9, value=f"{pair.get('sharpe', 0):.2f}").border = thin_border
    ws3.cell(row=row, column=10, value=f"{pair.get('calmar', 0):.2f}").border = thin_border
    row += 1

# ============= Sheet 4: Exit Reasons =============
ws4 = wb.create_sheet(title="Exit Reasons")

ws4['A1'] = 'Exit Reason Analysis'
ws4['A1'].font = title_font
ws4.merge_cells('A1:G1')

exit_reasons = nostalgia_data.get('exit_reason_summary', [])

headers = ['Exit Reason', 'Trades', 'Wins', 'Losses', 'Win Rate', 'Total Profit', 'Avg Profit %']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 4
for reason in exit_reasons:
    if reason.get('key') == 'TOTAL':
        continue
    ws4.cell(row=row, column=1, value=reason.get('key', '')).border = thin_border
    ws4.cell(row=row, column=2, value=reason.get('trades', 0)).border = thin_border
    ws4.cell(row=row, column=3, value=reason.get('wins', 0)).border = thin_border
    ws4.cell(row=row, column=4, value=reason.get('losses', 0)).border = thin_border
    ws4.cell(row=row, column=5, value=f"{reason.get('winrate', 0)*100:.1f}%").border = thin_border
    
    profit = reason.get('profit_total_abs', 0)
    profit_cell = ws4.cell(row=row, column=6, value=round(profit, 2))
    profit_cell.border = thin_border
    if profit > 0:
        profit_cell.fill = green_fill
    elif profit < 0:
        profit_cell.fill = red_fill
    
    ws4.cell(row=row, column=7, value=f"{reason.get('profit_mean_pct', 0):.2f}%").border = thin_border
    row += 1

# ============= Sheet 5: Daily PnL =============
ws5 = wb.create_sheet(title="Daily PnL")

ws5['A1'] = 'Daily Profit/Loss'
ws5['A1'].font = title_font
ws5.merge_cells('A1:C1')

daily_profit = nostalgia_data.get('daily_profit', [])

headers = ['Date', 'Profit (USDT)', 'Trades']
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 4
for day_data in daily_profit:
    date = day_data[0]
    profit = day_data[1]
    ws5.cell(row=row, column=1, value=date).border = thin_border
    
    profit_cell = ws5.cell(row=row, column=2, value=round(profit, 2))
    profit_cell.border = thin_border
    if profit > 0:
        profit_cell.fill = green_fill
    elif profit < 0:
        profit_cell.fill = red_fill
    
    row += 1

# ============= Sheet 6: Periodic Breakdown =============
ws6 = wb.create_sheet(title="Monthly Analysis")

ws6['A1'] = 'Monthly Breakdown'
ws6['A1'].font = title_font
ws6.merge_cells('A1:F1')

periodic = nostalgia_data.get('periodic_breakdown', {})
monthly = periodic.get('month', [])

headers = ['Month', 'Trades', 'Wins', 'Losses', 'Profit (USDT)', 'Profit Factor']
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 4
for month_data in monthly:
    ws6.cell(row=row, column=1, value=month_data.get('date', '')).border = thin_border
    ws6.cell(row=row, column=2, value=month_data.get('trades', 0)).border = thin_border
    ws6.cell(row=row, column=3, value=month_data.get('wins', 0)).border = thin_border
    ws6.cell(row=row, column=4, value=month_data.get('losses', 0)).border = thin_border
    
    profit = month_data.get('profit_abs', 0)
    profit_cell = ws6.cell(row=row, column=5, value=round(profit, 2))
    profit_cell.border = thin_border
    if profit > 0:
        profit_cell.fill = green_fill
    elif profit < 0:
        profit_cell.fill = red_fill
    
    ws6.cell(row=row, column=6, value=f"{month_data.get('profit_factor', 0):.2f}").border = thin_border
    row += 1

# ============= Sheet 7: Strategy Parameters =============
ws7 = wb.create_sheet(title="Strategy Parameters")

ws7['A1'] = 'Strategy Parameters'
ws7['A1'].font = title_font
ws7.merge_cells('A1:B1')

params = [
    ('Strategy Name', nostalgia_data.get('strategy_name', 'Nostalgia')),
    ('Timeframe', nostalgia_data.get('timeframe', '5m')),
    ('Trading Mode', nostalgia_data.get('trading_mode', 'futures')),
    ('Margin Mode', nostalgia_data.get('margin_mode', 'isolated')),
    ('Max Open Trades', nostalgia_data.get('max_open_trades', 3)),
    ('Stake Amount', f"{nostalgia_data.get('stake_amount', 100)} USDT"),
    ('Stoploss', f"{nostalgia_data.get('stoploss', -0.1)*100}%"),
    ('Trailing Stop', str(nostalgia_data.get('trailing_stop', True))),
    ('Trailing Stop Positive', f"{nostalgia_data.get('trailing_stop_positive', 0.01)*100}%"),
    ('Minimal ROI', str(nostalgia_data.get('minimal_roi', {}))),
    ('Use Exit Signal', str(nostalgia_data.get('use_exit_signal', True))),
    ('Exit Profit Only', str(nostalgia_data.get('exit_profit_only', True))),
]

for col, header in enumerate(['Parameter', 'Value'], 1):
    cell = ws7.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, (param, value) in enumerate(params, 4):
    ws7.cell(row=i, column=1, value=param).border = thin_border
    ws7.cell(row=i, column=2, value=str(value)).border = thin_border

# Save
output_path = '/home/neozh/freqtrade-strategies/user_data/backtest_results/nostalgia_complete_report.xlsx'
wb.save(output_path)
print(f"\n✅ Complete Excel report generated: {output_path}")
print(f"   - Total trades: {len(trades)}")
print(f"   - Sheets: Summary, All Trades, Pair Analysis, Exit Reasons, Daily PnL, Monthly Analysis, Strategy Parameters")