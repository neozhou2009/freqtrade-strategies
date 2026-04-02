#!/usr/bin/env python3
"""Create Excel report from Nostalgia backtest results"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "回测概览"

# Styles
title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=10)
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
alt_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
positive_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
negative_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:D1')
ws['A1'] = 'Nostalgia 策略回测报告'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Basic Info
ws['A3'] = '策略名称'
ws['B3'] = 'Nostalgia'
ws['C3'] = '报告生成时间'
ws['D3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

ws['A4'] = '回测时间范围'
ws['B4'] = '2025-01-02 至 2025-12-30 (362 天)'
ws['C4'] = '时间周期'
ws['D4'] = '5 分钟'

ws['A5'] = '交易所'
ws['B5'] = 'Binance (期货)'
ws['C5'] = '交易模式'
ws['D5'] = '逐仓模式'

ws['A6'] = '最大持仓数'
ws['B6'] = '3'
ws['C6'] = '单笔金额'
ws['D6'] = '100 USDT'

ws['A7'] = '初始资金'
ws['B7'] = '1000 USDT'
ws['C7'] = '最终资金'
ws['D7'] = '999.049 USDT'

# Performance Metrics
ws['A9'] = '业绩指标'
ws['A9'].font = header_font
ws['A9'].fill = header_fill
ws['B9'] = '数值'
ws['B9'].font = header_font
ws['B9'].fill = header_fill

metrics = [
    ('总交易次数', '6'),
    ('盈利交易', '3 (50.0%)'),
    ('亏损交易', '3 (50.0%)'),
    ('平均利润率', '-0.16%'),
    ('总利润 (USDT)', '-0.951'),
    ('总利润率', '-0.10%'),
    ('年化收益率 (CAGR)', '-0.10%'),
    ('夏普比率', '-0.04'),
    ('索提诺比率', '-0.10'),
    ('卡尔玛比率', '-1.18'),
    ('SQN', '-0.25'),
    ('盈利因子', '0.78'),
    ('期望值', '-0.16 (-0.11)'),
    ('平均每日利润', '-0.003 USDT'),
    ('最大回撤', '4.274 USDT (0.43%)'),
    ('回撤持续时间', '20 小时 55 分钟'),
]

for i, (label, value) in enumerate(metrics, 10):
    ws[f'A{i}'] = label
    ws[f'B{i}'] = value
    ws[f'A{i}'].font = normal_font
    ws[f'B{i}'].font = normal_font
    
    # Color coding for profit/loss
    if i == 13:  # Total profit
        try:
            if float(value.replace(' USDT', '').replace(',', '').replace('%', '')) > 0:
                ws[f'B{i}'].fill = positive_fill
            else:
                ws[f'B{i}'].fill = negative_fill
        except:
            pass

# Apply borders
for row in range(9, 10 + len(metrics)):
    for col in 'AB':
        ws[f'{col}{row}'].border = thin_border

# Pair Performance
ws2 = wb.create_sheet(title='交易对表现')
ws2['A1'] = '交易对表现统计'
ws2['A1'].font = title_font
ws2['A1'].fill = title_fill
ws2.merge_cells('A1:E1')
ws2.row_dimensions[1].height = 30

headers2 = ['交易对', '交易次数', '平均利润%', '总利润 (USDT)', '胜率']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

pair_data = [
    ('SOL/USDT:USDT', 3, 0.42, 1.276, '66.7%'),
    ('BTC/USDT:USDT', 2, -0.29, -0.599, '50.0%'),
    ('ETH/USDT:USDT', 1, -1.64, -1.628, '0.0%'),
    ('TOTAL', 6, -0.16, -0.951, '50.0%'),
]

for row, data in enumerate(pair_data, 4):
    for col, value in enumerate(data, 1):
        cell = ws2.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col == 5 and row == 4:  # SOL win rate
            cell.fill = positive_fill
        elif col == 5 and row == 6:  # ETH win rate 0%
            cell.fill = negative_fill

# Exit Reasons
ws3 = wb.create_sheet(title='退出原因分析')
ws3['A1'] = '退出原因统计'
ws3['A1'].font = title_font
ws3['A1'].fill = title_fill
ws3.merge_cells('A1:F1')
ws3.row_dimensions[1].height = 30

headers3 = ['退出原因', '次数', '平均利润%', '总利润 (USDT)', '平均持仓时间', '胜率']
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

exit_data = [
    ('signal_profit_u_1', 1, 2.07, 2.060, '0:05:00', '100%'),
    ('signal_profit_u_0', 1, 1.16, 1.152, '0:05:00', '100%'),
    ('signal_profit_u_e_1', 1, 0.11, 0.111, '0:05:00', '100%'),
    ('sell_signal_1', 3, -1.43, -4.274, '15:05:00', '0%'),
    ('TOTAL', 6, -0.16, -0.951, '7:35:00', '50.0%'),
]

for row, data in enumerate(exit_data, 4):
    for col, value in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if value == '100%':
            cell.fill = positive_fill
        elif value == '0%':
            cell.fill = negative_fill

# Strategy Parameters
ws4 = wb.create_sheet(title='策略参数')
ws4['A1'] = '策略主要参数'
ws4['A1'].font = title_font
ws4['A1'].fill = title_fill
ws4.merge_cells('A1:C1')
ws4.row_dimensions[1].height = 30

params = [
    ('INTERFACE_VERSION', '3'),
    ('timeframe', '5m'),
    ('minimal_roi', "{'0': 0.1, '30': 0.05, '60': 0.02}"),
    ('stoploss', '-0.1'),
    ('trailing_stop', 'True'),
    ('trailing_stop_positive', '0.01'),
    ('trailing_stop_positive_offset', '0.03'),
    ('trailing_only_offset_is_reached', 'True'),
    ('use_custom_stoploss', 'False'),
    ('use_exit_signal', 'True'),
    ('exit_profit_only', 'True'),
    ('ignore_roi_if_entry_signal', 'True'),
    ('max_open_trades', '3'),
    ('stake_amount', '100.0'),
    ('stake_currency', 'USDT'),
]

ws4['A3'] = '参数名'
ws4['B3'] = '参数值'
ws4['C3'] = '说明'
for col in range(1, 4):
    cell = ws4.cell(row=3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for row, (param, value) in enumerate(params, 4):
    ws4.cell(row=row, column=1, value=param).font = normal_font
    ws4.cell(row=row, column=2, value=value).font = normal_font
    ws4.cell(row=row, column=1).border = thin_border
    ws4.cell(row=row, column=2).border = thin_border

# Adjust column widths
for ws_sheet in wb.worksheets:
    for column in ws_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

# Save
output_path = '/home/neozh/freqtrade-strategies/user_data/backtest_results/nostalgia_backtest_report.xlsx'
wb.save(output_path)
print(f"✅ Excel 报告已生成：{output_path}")
